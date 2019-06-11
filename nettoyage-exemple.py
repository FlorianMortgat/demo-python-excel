#!/usr/bin/env python3
# -*- encoding: utf-8 -*-
"""
Script de démonstration de travail avec des données Excel. Nécessite la bibliothèque openpyxl :
    pip3 install openpyxl

Note : openpyxl est une bibliothèque libre (open-source) et écrite en python pur par
    Eric Gazoni
    Charlie Clark

- prend le fichier 'exemple-source.xlsx'
- analyse les données qu’il contient et tente de corriger
- exporte les données corrigées dans 'out/exemple-cible.xlsx'
- liste les ambiguités rencontrées dans une feuille séparée.

J’ai écrit les fonctions sans me soucier d’optimiser la performance ; pour des très gros
fichiers, il serait sans doute préférable de ne pas tout charger en mémoire et d’éviter
les structures de données « de confort ».

    Florian Mortgat, décembre 2018
"""

import os
import sys
import re

# j’utilise unicodedata pour asciifier des chaînes unicode
import unicodedata

# openpyxl est la bibliothèque de gestion du format xlsx
import openpyxl
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border
try: from openpyxl.styles import NamedStyle
except: from openpyxl.styles import Style as NamedStyle
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.table import Table, TableStyleInfo


# constantes liées aux fichiers de travail
IN_DIR = os.path.realpath(sys.path[0])
IN_FILEPATH = os.path.join(IN_DIR, 'exemple-source.xlsx')
OUT_DIR = os.path.join(IN_DIR, 'out')
OUT_FILEPATH = os.path.join(OUT_DIR, 'exemple-cible.xlsx')

# créer le répertoire de sortie si inexistant
if not os.path.exists(OUT_DIR):
    os.mkdir(OUT_DIR)

def demo_nettoyage():
    classeur_sale = openpyxl.load_workbook(IN_FILEPATH)
    donnees_sales = recuperer_donnees(classeur_sale)
    donnees_propres = analyser_et_corriger(*donnees_sales)
    classeur_propre = generer_classeur_propre(*donnees_propres)
    classeur_propre.save(OUT_FILEPATH)

def recuperer_donnees(wb):
    '''
    wb: Un objet 'Workbook' d’openpyxl (représentant un classeur Excel)
    retourne: 3 listes de lignes de cellules
    '''
    feuille_utilisateurs = wb['Utilisateurs']
    feuille_droits = wb['Droits']
    feuille_droits_utilisateurs = wb['Droits utilisateurs']

    utilisateurs = []
    # on spécifie min_row = 2 pour éviter d’inclure les cellules d’en-tête (première ligne).
    for user_ID_cell, nom_prenom_cell, *other_cells in feuille_utilisateurs.iter_rows(min_row = 2):
        # la propriété 'value' de l’objet cellule permet d’avoir le contenu.
        # pour les cellules de type numérique, le type de 'value' sera int, long ou float.
        # pour les cellules de type texte ou les formules, ce sera str (python 3) ou unicode (python 2).
        utilisateurs.append((user_ID_cell.value, nom_prenom_cell.value))

    droits = []
    for code_cell, droit_cell, *other_cells in feuille_droits.iter_rows(min_row=2):
        droits.append((code_cell.value, droit_cell.value))

    droits_utilisateurs = []
    for (
          nom_cell,
          prenom_cell,
          num_droit_cell,
          indice_droit_cell,
          *other_cells) in feuille_droits_utilisateurs.iter_rows(min_row = 2):
        droits_utilisateurs.append((nom_cell.value, prenom_cell.value, num_droit_cell.value, indice_droit_cell.value))
    return utilisateurs, droits, droits_utilisateurs

def analyser_et_corriger(utilisateurs, droits, droits_utilisateurs):
    '''
    utilisateurs: la liste de lignes de cellules de la feuille 'Utilisateurs' (sans la 1re ligne)
    droits: idem pour la feuille 'Droits'
    droits_utilisateurs: idem pour la feuille 'Droits utilisateurs'

    Homogénéise les données.
    '''
    ## 'Droits' est déjà OK
    droits_ok = droits

    ## on part de 'Droits utilisateurs' pour obtenir une liste normalisée des noms et prénoms
    droits_utilisateurs_ok = []
    noms_utilisateurs_ok = []
    for nom, prenom, num_droit, indice_droit in droits_utilisateurs:
        nom, prenom = map(supprimer_espaces_en_trop, (nom, prenom))
        code_droit = 'D{:03d}'.format(normaliser_nombre(num_droit))
        indice_droit = normaliser_nombre(indice_droit)
        droits_utilisateurs_ok.append((nom, prenom, code_droit, indice_droit))
        noms_utilisateurs_ok.append((nom, prenom))
    dict_noms_prenoms = map_nom_prenom_pour_comparaison(noms_utilisateurs_ok)

    ## dans 'Utilisateurs', il faut séparer les noms et les prénoms
    utilisateurs_ok = []
    for user_id, nom_prenom in utilisateurs:
        try:
            nom, prenom = demeler_nom_prenom(dict_noms_prenoms, nom_prenom)
        except KeyError:
            nom, prenom = nom_prenom, ''
        utilisateurs_ok.append((user_id, nom, prenom))

    ## on revient sur 'Droits utilisateurs' pour remplacer Nom et Prénom par un ID d’utilisateur
    uid_by_contact = {(nom, prenom): user_id for (user_id, nom, prenom) in utilisateurs_ok}
    droits_utilisateurs_ok = [
        (uid_by_contact[(nom, prenom)], code_droit, indice_droit)
        for (nom, prenom, code_droit, indice_droit) in droits_utilisateurs_ok
    ]
    return utilisateurs_ok, droits_ok, droits_utilisateurs_ok

def generer_classeur_propre(utilisateurs, droits, droits_utilisateurs):
    wb = openpyxl.Workbook()

    # 1) remplir les feuilles avec les données propres
    feuille_utilisateurs = wb.active
    feuille_utilisateurs.title = 'Utilisateurs'
    feuille_utilisateurs.append(('User ID', 'Nom', 'Prénom'))
    for ligne in utilisateurs:
        feuille_utilisateurs.append(ligne)

    feuille_droits = wb.create_sheet(title='Droits')
    feuille_droits.append(('Code', 'Droit'))
    for ligne in droits:
        feuille_droits.append(ligne)

    feuille_droits_utilisateurs = wb.create_sheet(title='Droits utilisateurs')
    feuille_droits_utilisateurs.append(('User ID', 'Code Droit', 'Indice d’utilisation du droit'))
    for ligne in droits_utilisateurs:
        feuille_droits_utilisateurs.append(ligne)

    # 2) créer la feuille 'Qui fait quoi' avec une formule pour indiquer qui a quel droit de façon lisible
    #    Note : Dans le format xlsx, les formules et les conventions suivent toujours le format anglais.
    #           C’est Excel qui se charge de les traduire selon la configuration.
    #           Du coup, dans openpyxl, tout est en anglais. 
    #           Si j’ouvre avec Excel (configuré en français) un classeur créé par openpyxl, je verrai bien
    #           les noms des formules en français, les arguments séparés par des points-virgules, les nombres
    #           avec des virgules et non des points, etc.

    feuille_qui_fait_quoi = wb.create_sheet(title='Qui fait quoi')
    feuille_qui_fait_quoi.append(('Qui', 'peut', 'Quoi',))

    plage_utilisateurs = '$A$2:$C${derniere_ligne}'.format(derniere_ligne = len(utilisateurs) + 1)
    plage_droits = '$A$2:$B${derniere_ligne}'.format(derniere_ligne = len(droits) + 1)

    wb.create_named_range('utilisateurs', feuille_utilisateurs, plage_utilisateurs)
    wb.create_named_range('droits', feuille_droits, plage_droits)   

    droits_utilisateurs_code_utilisateur = ''''Droits utilisateurs'!A{num_ligne:d}'''
    droits_utilisateurs_code_droit = ''''Droits utilisateurs'!B{num_ligne:d}'''
    formule_nom_par_id = '''VLOOKUP(%s,utilisateurs,2)'''%droits_utilisateurs_code_utilisateur
    formule_prenom_par_id = '''VLOOKUP(%s,utilisateurs,3)'''%droits_utilisateurs_code_utilisateur
    formule_prenom_nom_par_id = '''={} & " " & {}'''.format(formule_prenom_par_id, formule_nom_par_id)
    formule_droit_par_id = '''=LOWER(VLOOKUP(%s,droits,2))'''%droits_utilisateurs_code_droit
    for n in range(len(droits_utilisateurs)):
        num_ligne = n + 2 # +1 car une énumération python commence à 0 alors qu’Excel commence à 1, et +1 encore car saute l’en-tête
        aX = formule_prenom_nom_par_id.format(num_ligne = num_ligne)
        bX = 'peut'
        cX = formule_droit_par_id.format(num_ligne = num_ligne)
        feuille_qui_fait_quoi.append((aX, bX, cX))

    # 3) créer la feuille 'Cohérence' avec une formule pour détecter les doublons
    feuille_coherence = wb.create_sheet(title='Cohérence')
    feuille_coherence.append(('Indicateurs globaux', '', '', 'UID', 'unicité UID', 'Nom;Prénom', 'unicité Noms'))
    # 3.1) fusionner A1 et B1 dans la feuille de cohérence (cellule de titre)
    feuille_coherence.merge_cells('A1:B1')
    # 3.2) unicité de l’ID utilisateur et du prénom
    for n in range(len(utilisateurs)):
        num_ligne = n + 2
        valeurs = dict(
            D = '''='Utilisateurs'!A{X}'''.format(X=num_ligne),
            E = '=COUNTIF(cles_uid,D{X}) = 1'.format(X=num_ligne),
            F = '''='Utilisateurs'!B{X} & ";" & 'Utilisateurs'!C{X}'''.format(X=num_ligne),
            G = '=COUNTIF(cles_noms_prenoms,F{X}) = 1'.format(X=num_ligne)
        )
        for col in valeurs:
            feuille_coherence['{}{}'.format(col, num_ligne)].value = valeurs[col]
    # 3.3) création des plages nommées pour faciliter les formules
    plage_D = '$D$2:$D${max}'.format(max=len(utilisateurs)+1)
    plage_E = '$E$2:$E${max}'.format(max=len(utilisateurs)+1)
    plage_F = '$F$2:$F${max}'.format(max=len(utilisateurs)+1)
    plage_G = '$G$2:$G${max}'.format(max=len(utilisateurs)+1)
    wb.create_named_range('cles_uid', feuille_coherence, plage_D)
    wb.create_named_range('cles_uid_ok', feuille_coherence, plage_E)
    wb.create_named_range('cles_noms_prenoms', feuille_coherence, plage_F)
    wb.create_named_range('cles_noms_prenoms_ok', feuille_coherence, plage_G)

    # 3.4) ajout des formules de "cohérence globale"
    feuille_coherence['A2'].value = 'User ID uniques ?'
    feuille_coherence['B2'].value = '=COUNTIF(cles_uid_ok,FALSE) = 0'
    feuille_coherence['A3'].value = 'Noms+Prénoms uniques ?'
    feuille_coherence['B3'].value = '=COUNTIF(cles_noms_prenoms_ok,FALSE) = 0'
    plage_indicateurs_globaux = '$B$2:$B$3'

    # 4) après le contenu, la forme (styles, formatage, etc.) :
    # 4.1) ajuster la largeur des colonnes
    definir_largeur_colonnes(feuille_utilisateurs, dict(A=8, B=24, C=24))
    definir_largeur_colonnes(feuille_droits, dict(A=8, B=40))
    definir_largeur_colonnes(feuille_droits_utilisateurs, dict(A=8, B=12, C=27))
    definir_largeur_colonnes(feuille_qui_fait_quoi, dict(A=30, B=7, C=30))
    definir_largeur_colonnes(feuille_coherence, dict(A=30, B=10, C=4, D=9, E=17, F=31, G=19))
    # 4.2) en-têtes en gras
    def mettre_style_en_tete(cellule):
        cellule.style = STYLES.HEADER
    appliquer_a_plage(mettre_style_en_tete, feuille_utilisateurs['A1:C1'])
    appliquer_a_plage(mettre_style_en_tete, feuille_droits['A1:B1'])
    appliquer_a_plage(mettre_style_en_tete, feuille_droits_utilisateurs['A1:C1'])
    appliquer_a_plage(mettre_style_en_tete, feuille_qui_fait_quoi['A1:C1'])
    appliquer_a_plage(mettre_style_en_tete, feuille_coherence['A1:G1'])
    # 4.3) appliquer le format "pourcentage" aux nombres de la colonne C dans 'Droits utilisateurs'
    def format_pourcentage(cellule):
        cellule.style = 'Percent'
    derniere_ligne_de_droits_utilisateurs = len(droits_utilisateurs) + 1
    appliquer_a_plage(format_pourcentage, feuille_droits_utilisateurs['C2:C%d'%derniere_ligne_de_droits_utilisateurs])
    # 4.4) formatage conditionnel
    feuille_coherence.conditional_formatting.add('$E$2:$E${max}'.format(max=len(utilisateurs)+1), STYLES.HIGHLIGHT_FALSE_IN_RED)
    feuille_coherence.conditional_formatting.add('$G$2:$G${max}'.format(max=len(utilisateurs)+1), STYLES.HIGHLIGHT_FALSE_IN_RED)
    feuille_coherence.conditional_formatting.add(plage_indicateurs_globaux, STYLES.HIGHLIGHT_FALSE_IN_RED)
    feuille_coherence.conditional_formatting.add(plage_indicateurs_globaux, STYLES.HIGHLIGHT_TRUE_IN_GREEN)
    # 4.5) mise sous forme de tableau de la plage de vérifications de la feuille 'Cohérence'
    table_qui_fait_quoi = Table(displayName='Tableau_Qui_fait_quoi',
                                ref='$A$1:$C${max}'.format(max=len(droits_utilisateurs)+1),
                                tableStyleInfo=STYLES.PURPLE_TABLE)
    feuille_qui_fait_quoi.add_table(table_qui_fait_quoi)
    # 4.6) mise sous forme de tableau de la plage de vérifications de la feuille 'Cohérence'
    table_verifications = Table(displayName='Tableau_Vérifications',
                                ref='$D$1:$G${max}'.format(max=len(utilisateurs)+1),
                                tableStyleInfo=STYLES.BLUE_TABLE)
    feuille_coherence.add_table(table_verifications)
    wb.active = feuille_coherence
    return wb

def supprimer_espaces_en_trop(valeur_brute):
    return re.sub('  +', ' ', valeur_brute).strip()

def normaliser_nombre(valeur_brute):
    t = type(valeur_brute)
    if t not in (int, float, str):
        raise TypeError('valeur_brute should be either int, float or str')
    if t is str:
        valeur_brute = valeur_brute.replace(' ', '')
        valeur_brute = valeur_brute.replace(',', '.')

        # gérer le cas des pourcentages
        multiplier_par = 1
        if valeur_brute.endswith('%'):
            multiplier_par = 0.01
            valeur_brute = valeur_brute[:-1]
        if re.match('\d+$', valeur_brute):
            return multiplier_par * int(valeur_brute)
        return multiplier_par * float(valeur_brute)
    elif t in (int, float):
        return valeur_brute

def asciifier(chaine):
    '''
    Supprime les diacritiques (accents, points, cédilles, etc.) et défait les ligatures.
    Passe le tout en minuscules.
    Note : cette fonction ne latinisera pas les systèmes d’écriture non latins
    (cyrillique, CJK, etc.) – ce serait beaucoup plus compliqué
    '''
    return (
        unicodedata.normalize(
            'NFKD',
            (
                chaine.lower()
                .replace('œ', 'oe') # ligatures et autres non gérés par unicodedata
                .replace('æ', 'ae')
                .replace('ĳ', 'ij')
                .replace('ß', 'ss')
                .replace('ǉ', 'lj')
                .replace('ǌ', 'nj')
                .replace('ﬆ', 'st')
                .replace('ﬅ', 'ft')
                .replace('ǳ', 'dz')
                .replace('ﬀ', 'ff')
                .replace('ſ', 's') # s long
                .replace(' ', ' ') # espace insécable fine
                .replace(' ', ' ') # espace insécable normale
                .replace('\t', ' ') # tabulation
                .replace(' ', '')
                .replace('—', '') # tiret cadratin
                .replace('–', '') # tiret
                .replace('-', '') # signe moins
                .replace('’', '') # apostrophe courbe
                .replace("'", '') # apostrophe droite
            ))
        .encode('ascii', 'ignore') # supprime les caractères non ascii
        .decode('ascii') # retransforme l’objet 'bytes' en 'str'
    )

def map_nom_prenom_pour_comparaison(liste_noms_prenoms):
    return {asciifier(nom + prenom): (nom, prenom) for (nom, prenom) in set(liste_noms_prenoms)}

def demeler_nom_prenom(dict_reference, nom_prenom):
    cle = asciifier(nom_prenom)
    if cle in dict_reference:
        return dict_reference[cle]
    else:
        raise KeyError("'nom_prenom' ({}) introuvable dans dict_reference.".format(repr(nom_prenom)))

def definir_largeur_colonnes(feuille, dict_colonnes):
    for lettre_col, largeur in dict_colonnes.items():
        feuille.column_dimensions[lettre_col].width = largeur

def appliquer_a_plage(predicat, iter_plage):
    for ligne in iter_plage:
        for cellule in ligne:
            predicat(cellule)

class STYLES:
    '''
    Classe de stockage de mes styles Excel.
    '''
    # Création de styles Excel qu’on pourra appliquer à des cellules
    BIG_BOLD = Font(bold=True, size=12)
    UNDERLINE = Font(underline="single")

    CENTERED = Alignment(horizontal='center')
    WRAP = Alignment(wrap_text=True)

    HEADER = NamedStyle(name='header_cell', font=BIG_BOLD, alignment=CENTERED)

    # styles pour le formatage conditionnel
    GREEN_BG = DifferentialStyle(
        font=Font(color='FF006100'),
        fill=PatternFill(bgColor='FFC6EFCE'))
    RED_BG = DifferentialStyle(
        font=Font(color='FF9C0006'),
        fill=PatternFill(bgColor='FFFFC7CE'))

    # Règles de formatage conditionnel : VRAI = vert, FAUX = rouge.
    HIGHLIGHT_FALSE_IN_RED = Rule(
        type='cellIs',
        dxf=RED_BG,
        operator='equal',
        formula=['FALSE'])
    HIGHLIGHT_TRUE_IN_GREEN = Rule(
        type='cellIs',
        dxf=GREEN_BG,
        operator='equal',
        formula=['TRUE'])

    # Ces 2 styles de tableau (ainsi que d’autres) sont incorporés à Excel par défaut.
    PURPLE_TABLE = TableStyleInfo(
        name="TableStyleMedium5",
        showRowStripes=True,
        showColumnStripes=False)
    BLUE_TABLE = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False)

def main():
    demo_nettoyage()

if __name__ == '__main__':
    main()
